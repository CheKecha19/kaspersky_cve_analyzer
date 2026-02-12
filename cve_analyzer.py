import pandas as pd
import requests
import time
import json
import os
import sys
import random
import logging
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import urllib3
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import argparse
from tenacity import (
    retry,
    stop_after_attempt,
    wait_exponential,
    wait_random,
    wait_combine,
    retry_if_exception_type,
    before_sleep_log
)

# ==============================
# НАСТРОЙКИ ЛОГИРОВАНИЯ
# ==============================
LOG_LEVEL = "INFO"

# Настройка логирования
log_formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')

file_handler = logging.FileHandler('cve_analyzer.log', encoding='utf-8')
file_handler.setFormatter(log_formatter)
file_handler.setLevel(getattr(logging, LOG_LEVEL))

console_handler = logging.StreamHandler()
console_handler.setFormatter(log_formatter)
console_handler.setLevel(logging.WARNING)

logger = logging.getLogger()
logger.setLevel(getattr(logging, LOG_LEVEL))
logger.addHandler(file_handler)
logger.addHandler(console_handler)

api_logger = logging.getLogger('NVD_API')
exploit_logger = logging.getLogger('EXPLOIT_CHECK')
main_logger = logging.getLogger('MAIN')
file_logger = logging.getLogger('FILE_HANDLING')

# ==============================
# КОНЕЦ НАСТРОЕК ЛОГИРОВАНИЯ
# ==============================

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Чтение API ключа из переменных окружения
NVD_API_KEY = os.getenv('NVD_API_KEY', '93d9dc54-a6d1-4d16-a62c-c7cfa5351bca')

# Глобальные настройки задержек
NVD_DELAY = 1.5
EXPLOIT_DELAY = 2
MAX_RETRIES = 5

MODES = {
    '1': 'third-party',
    '2': 'microsoft', 
    '3': 'all'
}

# Файл для сохранения прогресса
PROGRESS_FILE = "cve_analyzer_progress.json"

# Настройки Tenacity для повторных попыток
NVD_RETRY_CONFIG = {
    "stop": stop_after_attempt(MAX_RETRIES),
    "wait": wait_combine(
        wait_exponential(multiplier=1, min=4, max=60),
        wait_random(0, 2)
    ),
    "retry": retry_if_exception_type((requests.exceptions.RequestException, urllib3.exceptions.HTTPError)),
    "before_sleep": before_sleep_log(api_logger, logging.WARNING),
    "reraise": False
}

def severity_to_number(severity):
    """Преобразует текстовую оценку критичности в числовое значение для сравнения"""
    if not severity or severity == 'N/A':
        return 0
        
    severity_map = {
        'None': 0,
        'Low': 1,
        'Medium': 2, 
        'High': 3,
        'Critical': 4,
    }
    return severity_map.get(severity, 0)

def number_to_severity(number):
    """Преобразует числовое значение обратно в текстовую оценку критичности"""
    severity_map = {
        0: 'None',
        1: 'Low', 
        2: 'Medium',
        3: 'High',
        4: 'Critical'
    }
    return severity_map.get(number, 'N/A')

def get_highest_severity(severity1, severity2):
    """Возвращает наивысшую оценку критичности из двух"""
    num1 = severity_to_number(severity1)
    num2 = severity_to_number(severity2)
    
    highest_num = max(num1, num2)
    return number_to_severity(highest_num)

def create_session():
    """Создает сессию с повторными попытками и отключенной проверкой SSL"""
    main_logger.debug("Создание HTTP сессии с повторными попытками")
    session = requests.Session()
    
    retry_strategy = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET"],
        respect_retry_after_header=True
    )
    
    adapter = HTTPAdapter(max_retries=retry_strategy, pool_connections=10, pool_maxsize=10)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    session.verify = False
    
    session.timeout = (30, 60)
    
    main_logger.debug("HTTP сессия создана успешно")
    return session

def find_column_name(df, possible_names):
    """Находит имя столбца в DataFrame по списку возможных имен"""
    for name in possible_names:
        if name in df.columns:
            return name
    return None

def find_details_sheet(file_path):
    """Автоматически находит лист с CVE данными в файле Excel"""
    file_logger.info(f"Поиск листа с CVE данными в файле: {file_path}")
    try:
        excel_file = pd.ExcelFile(file_path)
        sheet_names = excel_file.sheet_names
        
        file_logger.debug(f"Найдены листы в файле: {sheet_names}")
        
        for i, sheet_name in enumerate(sheet_names):
            try:
                test_df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=10)
                
                cve_column = find_column_name(test_df, [
                    'CVE ID', 'CVE_ID', 'CVE-ID', 'CVE', 
                    'Vulnerability entry ID', 'Vulnerability ID'
                ])
                
                severity_column = find_column_name(test_df, [
                    'Severity level', 'Severity', 'Level', 'Criticality'
                ])
                
                if cve_column and severity_column:
                    file_logger.info(f"Найден лист с CVE данными: '{sheet_name}' (индекс {i})")
                    return sheet_name, i
                    
            except Exception as e:
                file_logger.debug(f"Ошибка при проверке листа {sheet_name}: {e}")
                continue
        
        if sheet_names:
            file_logger.warning(f"Используем первый лист: '{sheet_names[0]}'")
            return sheet_names[0], 0
        else:
            file_logger.error("В файле нет листов")
            return None, None
        
    except Exception as e:
        file_logger.error(f"Ошибка при поиске листа: {e}", exc_info=True)
        return None, None

@retry(**NVD_RETRY_CONFIG)
def get_cvss_score_with_retry(cve_id, session):
    """Получает CVSS оценку для CVE через API NVD с использованием Tenacity для повторных попыток"""
    api_logger.debug(f"Запрос данных для {cve_id}")
    url = f"https://services.nvd.nist.gov/rest/json/cves/2.0?cveId={cve_id}"
    
    headers = {}
    if NVD_API_KEY:
        headers["apiKey"] = NVD_API_KEY
    
    start_time = time.time()
    response = session.get(url, headers=headers, timeout=(30, 60))
    response_time = time.time() - start_time
    
    api_logger.debug(f"Ответ от NVD API для {cve_id}: статус {response.status_code}, время: {response_time:.2f}с")
    
    if response.status_code == 403:
        api_logger.error(f"Получен статус 403 для {cve_id}. Проверьте API ключ.")
        return "Error: 403 Forbidden", "API Error", "N/A", "N/A"
    elif response.status_code == 429:
        api_logger.warning(f"Превышен лимит запросов (429) для {cve_id}")
        return "Error: 429 Rate Limited", "API Error", "N/A", "N/A"
    elif response.status_code != 200:
        api_logger.warning(f"Некорректный статус ответа для {cve_id}: {response.status_code}")
    
    response.raise_for_status()
    
    data = response.json()
    
    if 'vulnerabilities' in data and len(data['vulnerabilities']) > 0:
        vulnerability = data['vulnerabilities'][0]['cve']
        
        published_date = vulnerability.get('published', 'N/A')
        
        if 'cvssMetricV31' in vulnerability['metrics']:
            cvss_data = vulnerability['metrics']['cvssMetricV31'][0]['cvssData']
            base_score = cvss_data['baseScore']
            version = cvss_data['version']
            api_logger.debug(f"{cve_id}: CVSS {version} оценка = {base_score}")
        elif 'cvssMetricV30' in vulnerability['metrics']:
            cvss_data = vulnerability['metrics']['cvssMetricV30'][0]['cvssData']
            base_score = cvss_data['baseScore']
            version = cvss_data['version']
            api_logger.debug(f"{cve_id}: CVSS {version} оценка = {base_score}")
        elif 'cvssMetricV2' in vulnerability['metrics']:
            cvss_data = vulnerability['metrics']['cvssMetricV2'][0]['cvssData']
            base_score = cvss_data['baseScore']
            version = "2.0"
            api_logger.debug(f"{cve_id}: CVSS {version} оценка = {base_score}")
        else:
            base_score = "N/A"
            version = "No CVSS data"
            api_logger.debug(f"{cve_id}: данные CVSS не найдены")
        
        if isinstance(base_score, (int, float)):
            if base_score >= 9.0:
                severity = "Critical"
            elif base_score >= 7.0:
                severity = "High"
            elif base_score >= 4.0:
                severity = "Medium"
            elif base_score >= 0.1:
                severity = "Low"
            else:
                severity = "None"
        else:
            severity = "N/A"
            
        api_logger.info(f"{cve_id}: оценка {base_score}, уровень {severity}, версия {version}")
        return base_score, version, severity, published_date
    else:
        api_logger.warning(f"{cve_id}: уязвимость не найдена в NVD")
        return "Not Found", "CVE not found", "N/A", "N/A"

def get_cvss_score(cve_id, session):
    """Обертка вокруг get_cvss_score_with_retry с обработкой исключений"""
    try:
        return get_cvss_score_with_retry(cve_id, session)
    except requests.exceptions.RequestException as e:
        api_logger.error(f"Ошибка запроса для {cve_id} после {MAX_RETRIES} попыток: {str(e)}")
        return f"Error: {str(e)}", "API Error", "N/A", "N/A"
    except (KeyError, IndexError, json.JSONDecodeError) as e:
        api_logger.error(f"Ошибка парсинга данных для {cve_id}: {str(e)}", exc_info=True)
        return f"Parse Error: {str(e)}", "Data Parse Error", "N/A", "N/A"
    except Exception as e:
        api_logger.error(f"Неожиданная ошибка для {cve_id}: {str(e)}", exc_info=True)
        return f"Unexpected Error: {str(e)}", "System Error", "N/A", "N/A"

@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=2, max=10),
    retry=retry_if_exception_type(requests.exceptions.RequestException)
)
def check_exploit_availability_with_retry(cve_id, session):
    """Проверяет наличие эксплойтов для CVE с использованием Tenacity"""
    exploit_logger.debug(f"Проверка эксплойтов для {cve_id}")
    sources = [
        f"https://www.exploit-db.com/search?cve={cve_id}",
        f"https://github.com/search?q={cve_id}+exploit&type=repositories",
        f"https://packetstormsecurity.com/search/?q={cve_id}",
        f"https://0day.today/search?search_request={cve_id}"
    ]
    
    source_names = ["exploit-db.com", "github.com", "packetstormsecurity.com", "0day.today"]
    exploit_sources = []
    
    for i, source in enumerate(sources):
        try:
            exploit_logger.debug(f"Проверка {source_names[i]} для {cve_id}")
            response = session.get(source, timeout=15, verify=False)
            if response.status_code == 200:
                if any(keyword in response.text.lower() for keyword in ['exploit', 'poc', 'proof of concept', 'code execution']):
                    exploit_sources.append(source_names[i])
                    exploit_logger.debug(f"Найден эксплойт для {cve_id} в {source_names[i]}")
            else:
                exploit_logger.debug(f"Статус {response.status_code} для {source_names[i]}")
        except Exception as e:
            exploit_logger.debug(f"Ошибка при проверке {source_names[i]}: {e}")
            continue
        
        time.sleep(0.5 + random.uniform(0, 0.5))
    
    result = f"Найдены в: {', '.join(exploit_sources)}" if exploit_sources else "Не найдено подтверждений о наличии публичных эксплойтов"
    exploit_logger.info(f"Результат проверки эксплойтов для {cve_id}: {result}")
    return result

def check_exploit_availability(cve_id, session):
    """Обертка вокруг check_exploit_availability_with_retry с обработкой исключений"""
    try:
        return check_exploit_availability_with_retry(cve_id, session)
    except Exception as e:
        exploit_logger.error(f"Ошибка при проверке эксплойтов для {cve_id}: {e}")
        return f"Ошибка проверки: {str(e)}"

def print_progress_bar(iteration, total, prefix='', suffix='', decimals=1, length=50, fill='█'):
    """Выводит прогресс-бар в консоль"""
    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filled_length = int(length * iteration // total)
    bar = fill * filled_length + '-' * (length - filled_length)
    print(f'\r{prefix} |{bar}| {percent}% {suffix}', end='\r')
    if iteration == total:
        print()

def save_progress(mode, processed_cves, stats, current_index, total_cves, output_file):
    """Сохраняет прогресс обработки в файл"""
    progress_data = {
        'mode': mode,
        'processed_cves': list(processed_cves),
        'stats': stats,
        'current_index': current_index,
        'total_cves': total_cves,
        'output_file': output_file,
        'timestamp': datetime.now().isoformat()
    }
    
    try:
        with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
            json.dump(progress_data, f, indent=2, ensure_ascii=False)
        main_logger.debug(f"Прогресс сохранен: обработано {len(processed_cves)} CVE")
    except Exception as e:
        main_logger.error(f"Ошибка при сохранении прогресса: {e}")

def load_progress():
    """Загружает прогресс обработки из файла"""
    if not os.path.exists(PROGRESS_FILE):
        return None
    
    try:
        with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
            progress_data = json.load(f)
        main_logger.info(f"Загружен прогресс: {len(progress_data['processed_cves'])} обработанных CVE")
        return progress_data
    except Exception as e:
        main_logger.error(f"Ошибка при загрузке прогресса: {e}")
        return None

def initialize_excel_output(output_file):
    """Инициализирует Excel файл с заголовками"""
    main_logger.info(f"Инициализация выходного файла: {output_file}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'CVE Report'
    
    headers = [
        'Узел',
        'CVE-идентификатор', 
        'Дата публикации информации об уязвимости',
        'Оценка уязвимости CVSSv3',
        'Оценка критичности согласно CVSSv3',
        'Оценка критичности согласно Kaspersky',
        'Оценка критичности согласно внутренней методики',
        'Наличие exploit/workaround',
        'Дата обнаружения',
        'Дата устранения (фактическая или запланированная)',
        'Комментарий'
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    column_widths = {
        'A': 25, 'B': 20, 'C': 25, 'D': 15, 'E': 25,
        'F': 30, 'G': 35, 'H': 35, 'I': 20, 'J': 30, 'K': 40
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    wb.save(output_file)
    main_logger.info("Выходной файл инициализирован успешно")
    return len(headers)

def append_to_excel(output_file, row_data, row_number):
    """Добавляет строку данных в Excel файл"""
    try:
        wb = load_workbook(output_file)
        ws = wb.active
        
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_number, column=col, value=value)
        
        wb.save(output_file)
        return True
    except Exception as e:
        main_logger.error(f"Ошибка при записи в файл {output_file}: {e}")
        return False

def apply_color_formatting(output_file):
    """Применяет цветовое форматирование к Excel файлу"""
    main_logger.debug(f"Применение цветового форматирования к файлу: {output_file}")
    try:
        wb = load_workbook(output_file)
        ws = wb.active
        
        critical_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        high_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        medium_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        low_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        na_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        formatted_cells = 0
        for row in range(2, ws.max_row + 1):
            score_cell = ws.cell(row=row, column=4)
            
            try:
                if isinstance(score_cell.value, (int, float)):
                    score = float(score_cell.value)
                    
                    if score >= 9.0:
                        score_cell.fill = critical_fill
                    elif score >= 7.0:
                        score_cell.fill = high_fill
                    elif score >= 4.0:
                        score_cell.fill = medium_fill
                    elif score >= 0.1:
                        score_cell.fill = low_fill
                    formatted_cells += 1
                else:
                    score_cell.fill = na_fill
                    formatted_cells += 1
                    
            except (ValueError, TypeError):
                score_cell.fill = na_fill
                formatted_cells += 1
        
        wb.save(output_file)
        main_logger.info(f"Цветовое форматирование применено к {formatted_cells} ячейкам")
    except Exception as e:
        main_logger.error(f"Ошибка при применении форматирования: {e}")

def get_output_filename(input_file, mode):
    """Генерирует имя выходного файла на основе имени входного файла и режима"""
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    return f"{base_name}_{mode}.xlsx"

def process_cve_list(file_path, mode='third-party', check_exploits=True, resume=False):
    """Обрабатывает файл Excel и получает CVSS оценки для CVE с пошаговой записью"""
    main_logger.info(f"Начало обработки файла: {file_path} в режиме: {mode}, проверка эксплойтов: {check_exploits}")
    
    progress_data = None
    processed_cves = set()
    if resume:
        progress_data = load_progress()
        if progress_data and progress_data.get('mode') == mode:
            processed_cves = set(progress_data.get('processed_cves', []))
            main_logger.info(f"Найдено {len(processed_cves)} ранее обработанных CVE")
    
    details_sheet_name, details_sheet_index = find_details_sheet(file_path)
    
    if details_sheet_name is None:
        main_logger.error("Не удалось найти подходящий лист с CVE данными")
        return {}, []
    
    try:
        main_logger.info(f"Чтение листа: '{details_sheet_name}'")
        df = pd.read_excel(file_path, sheet_name=details_sheet_name)
        main_logger.info(f"Успешно прочитан лист '{details_sheet_name}', колонки: {list(df.columns)}")
    except Exception as e:
        main_logger.error(f"Ошибка при чтении листа '{details_sheet_name}': {e}", exc_info=True)
        return {}, []
    
    cve_column = find_column_name(df, [
        'CVE ID', 'CVE_ID', 'CVE-ID', 'CVE', 
        'Vulnerability entry ID', 'Vulnerability ID'
    ])
    vendor_column = find_column_name(df, ['Vendor', 'Vendor name', 'Производитель'])
    device_column = find_column_name(df, ['Device', 'Устройство', 'Host', 'Хост'])
    app_column = find_column_name(df, ['Application name', 'Application', 'Приложение'])
    severity_column = find_column_name(df, ['Severity level', 'Severity', 'Уровень критичности', 'Критичность'])
    
    if not cve_column:
        main_logger.error("Столбец с CVE ID не найден в файле")
        main_logger.error(f"Доступные колонки: {list(df.columns)}")
        return {}, []
    
    main_logger.info(f"Используемые столбцы: CVE={cve_column}, Vendor={vendor_column}, Device={device_column}, App={app_column}, Severity={severity_column}")
    
    if vendor_column:
        initial_count = len(df)
        
        if mode == 'third-party':
            df = df[df[vendor_column] != 'Microsoft']
            filtered_count = initial_count - len(df)
            main_logger.info(f"Режим 'third-party': отфильтровано записей Microsoft: {filtered_count}")
        elif mode == 'microsoft':
            df = df[df[vendor_column] == 'Microsoft']
            filtered_count = initial_count - len(df)
            main_logger.info(f"Режим 'microsoft': отфильтровано записей не-Microsoft: {filtered_count}")
        else:
            main_logger.info(f"Режим 'all': все записи сохранены, без фильтрации")
            filtered_count = 0
    
    cve_data = {}
    
    main_logger.info("Извлечение уникальных CVE из файла с оценками Kaspersky")
    for index, row in df.iterrows():
        cve_id = str(row[cve_column]) if pd.notna(row[cve_column]) else ""
        if cve_id.startswith('CVE-'):
            if resume and cve_id in processed_cves:
                continue
                
            device = str(row[device_column]) if device_column and pd.notna(row.get(device_column, '')) else ""
            application = str(row[app_column]) if app_column and pd.notna(row.get(app_column, '')) else ""
            
            kaspersky_severity = "N/A"
            if severity_column and pd.notna(row.get(severity_column)):
                kaspersky_severity = str(row[severity_column])
            
            if cve_id not in cve_data:
                cve_data[cve_id] = {
                    'devices': set([device]) if device else set(),
                    'applications': set([application]) if application else set(),
                    'kaspersky_severity': kaspersky_severity
                }
            else:
                if device:
                    cve_data[cve_id]['devices'].add(device)
                if application:
                    cve_data[cve_id]['applications'].add(application)
    
    cve_list = list(cve_data.keys())
    cve_list.sort()
    
    main_logger.info(f"Найдено {len(cve_list)} уникальных CVE для обработки (режим: {mode})")
    
    # Новое именование файлов: <имя_файла>_<режим>.xlsx
    output_dir = os.path.dirname(file_path)
    output_file = os.path.join(output_dir, get_output_filename(file_path, mode))
    
    if not os.path.exists(output_file):
        num_columns = initialize_excel_output(output_file)
        current_row = 2
    else:
        try:
            wb = load_workbook(output_file)
            ws = wb.active
            current_row = ws.max_row + 1
            main_logger.info(f"Продолжение записи в существующий файл с строки {current_row}")
        except Exception as e:
            main_logger.error(f"Ошибка при чтении существующего файла: {e}")
            num_columns = initialize_excel_output(output_file)
            current_row = 2
    
    session = create_session()
    
    stats = {
        'critical': 0,
        'high': 0,
        'medium': 0,
        'low': 0,
        'errors': 0,
        'not_found': 0,
        'rate_limited': 0
    }
    
    if resume and progress_data:
        stats = progress_data.get('stats', stats)
        main_logger.info(f"Восстановлена статистика: {stats}")
    
    top_cves = []
    today_date = datetime.now().strftime('%Y-%m-%d 00:00:00')
    
    main_logger.info("Начало обработки CVE через NVD API с пошаговой записью")
    
    for i, cve_id in enumerate(cve_list):
        base_score, version, cvss_severity, published_date = get_cvss_score(cve_id, session)
        
        exploit_info = "Проверка отключена"
        if check_exploits:
            time.sleep(EXPLOIT_DELAY + random.uniform(0, 0.5))
            exploit_info = check_exploit_availability(cve_id, session)
        
        kaspersky_severity = cve_data[cve_id].get('kaspersky_severity', 'N/A')
        
        if cvss_severity == "N/A" or cvss_severity == "CVE not found":
            combined_severity = kaspersky_severity
        else:
            combined_severity = get_highest_severity(cvss_severity, kaspersky_severity)
        
        severity_num = severity_to_number(combined_severity)
        if severity_num == 4:
            stats['critical'] += 1
        elif severity_num == 3:
            stats['high'] += 1
        elif severity_num == 2:
            stats['medium'] += 1
        elif severity_num == 1:
            stats['low'] += 1
        
        if isinstance(base_score, (int, float)):
            top_cves.append((cve_id, base_score, combined_severity))
            top_cves.sort(key=lambda x: x[1], reverse=True)
            if len(top_cves) > 5:
                top_cves = top_cves[:5]
        elif "403 Forbidden" in str(base_score):
            stats['rate_limited'] += 1
        elif "Not Found" in str(base_score):
            stats['not_found'] += 1
        else:
            stats['errors'] += 1
        
        if published_date != 'N/A' and published_date != 'N/A':
            try:
                published_date = pd.to_datetime(published_date).strftime('%Y-%m-%d 00:00:00')
            except Exception as e:
                main_logger.debug(f"Ошибка форматирования даты для {cve_id}: {e}")
                published_date = 'N/A'
        
        devices = ', '.join(cve_data[cve_id]['devices']) if cve_data[cve_id]['devices'] else ''
        applications = ', '.join(cve_data[cve_id]['applications']) if cve_data[cve_id]['applications'] else ''
        
        row_data = [
            devices,
            cve_id,
            published_date,
            base_score if isinstance(base_score, (int, float)) else 'N/A',
            cvss_severity,
            kaspersky_severity,
            combined_severity,
            exploit_info,
            today_date,
            '',
            applications
        ]
        
        success = append_to_excel(output_file, row_data, current_row)
        if not success:
            main_logger.error(f"Не удалось записать данные для {cve_id}")
        
        current_row += 1
        
        processed_cves.add(cve_id)
        if i % 10 == 0:
            save_progress(mode, processed_cves, stats, i, len(cve_list), output_file)
        
        progress_prefix = f"Режим {mode}: {stats['critical']} критич, {stats['high']} высок, {stats['medium']} средн, {stats['low']} низк, {stats['errors']} ошибок"
        print_progress_bar(i + 1, len(cve_list), prefix=progress_prefix, suffix=f'Обработано: {i+1}/{len(cve_list)}')
        
        time.sleep(NVD_DELAY + random.uniform(0, 0.3))
    
    print()
    main_logger.info("Обработка всех CVE завершена")
    
    if os.path.exists(PROGRESS_FILE):
        os.remove(PROGRESS_FILE)
        main_logger.info("Файл прогресса удален")
    
    apply_color_formatting(output_file)
    
    print(f"Результаты сохранены в файл: {output_file}")
    return stats, top_cves

def get_file_path():
    """Запрашивает путь к файлу у пользователя"""
    default_path = r"C:\Users\cu-nazarov-na\Desktop\kasper_cve\example.xlsx"
    
    print("Введите путь к файлу Excel с CVE данными:")
    print(f"Нажмите Enter для использования пути по умолчанию: {default_path}")
    
    user_input = input("Путь к файлу: ").strip()
    
    if user_input == "":
        file_path = default_path
    else:
        file_path = user_input
    
    if file_path.startswith('"') and file_path.endswith('"'):
        file_path = file_path[1:-1]
    
    return file_path

def get_work_mode():
    """Запрашивает у пользователя режим работы"""
    print("\nВыберите режим работы:")
    print("1 - Third-party (исключить Microsoft) - по умолчанию")
    print("2 - Microsoft (только Microsoft)")
    print("3 - All (все CVE)")
    
    user_input = input("Введите номер режима (1/2/3): ").strip()
    
    if user_input in MODES:
        mode = MODES[user_input]
        print(f"Выбран режим: {mode}")
    else:
        mode = 'third-party'
        print(f"Неверный ввод, используется режим по умолчанию: {mode}")
    
    return mode

def ask_resume():
    """Спрашивает пользователя о возобновлении работы"""
    if os.path.exists(PROGRESS_FILE):
        print("\nОбнаружен файл прогресса предыдущего запуска.")
        response = input("Хотите возобновить обработку? (y/n): ").strip().lower()
        return response in ['y', 'yes', 'д', 'да']
    return False

def main():
    parser = argparse.ArgumentParser(
        description='Анализатор CVE уязвимостей из отчетов Kaspersky Security Center',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Примеры использования:
  python cve_analyzer.py                                    # Запуск с настройками по умолчанию
  python cve_analyzer.py --no-exploits                     # Без проверки эксплойтов
  python cve_analyzer.py --no-resume                       # Начать заново без возобновления
  python cve_analyzer.py --no-exploits --no-resume         # Полный перезапуск без проверки эксплойтов

Режимы работы:
  1 - Third-party: исключить уязвимости Microsoft
  2 - Microsoft: только уязвимости Microsoft  
  3 - All: все уязвимости без фильтрации
        '''
    )
    parser.add_argument('--no-exploits', action='store_true', 
                       help='Отключить проверку эксплойтов (по умолчанию проверка включена)')
    parser.add_argument('--no-resume', action='store_true', 
                       help='Не возобновлять предыдущую обработку (начать заново)')
    
    args = parser.parse_args()
    
    # Проверка эксплойтов включена по умолчанию, флаг --no-exploits отключает ее
    check_exploits = not args.no_exploits
    
    main_logger.info("Запуск CVE анализатора")
    
    file_path = get_file_path()
    mode = get_work_mode()
    
    resume = False
    if not args.no_resume:
        resume = ask_resume()
    
    if not os.path.exists(file_path):
        main_logger.error(f"Файл не найден: {file_path}")
        print(f"Файл не найден: {file_path}")
        return
    
    print("Начинаем анализ CVE уязвимостей...")
    print(f"Исходный файл: {file_path}")
    print(f"Режим работы: {mode}")
    print(f"Проверка эксплойтов: {'ВКЛЮЧЕНА' if check_exploits else 'ОТКЛЮЧЕНА'}")
    print(f"Возобновление работы: {'ДА' if resume else 'НЕТ'}")
    print(f"API ключ: {'УСТАНОВЛЕН' if NVD_API_KEY else 'НЕ НАЙДЕН'}")
    print(f"Уровень логирования: {LOG_LEVEL}")
    print(f"Настройки задержек: NVD={NVD_DELAY}с, Exploit={EXPLOIT_DELAY}с")
    print(f"Максимальное количество попыток: {MAX_RETRIES}")
    print("=" * 60)
    
    main_logger.info(f"Настройки: режим={mode}, check_exploits={check_exploits}, resume={resume}")
    
    try:
        stats, top_cves = process_cve_list(file_path, mode, check_exploits, resume)
    except Exception as e:
        main_logger.critical(f"Критическая ошибка при выполнении анализа: {e}", exc_info=True)
        print(f"Произошла критическая ошибка: {e}")
        print("Подробности смотрите в файле лога cve_analyzer.log")
        return
    
    print("\n" + "=" * 60)
    print("ИТОГОВАЯ СТАТИСТИКА АНАЛИЗА")
    print("=" * 60)
    print(f"Режим работы: {mode}")
    print(f"Всего обработано CVE: {sum(stats.values())}")
    print(f"Критические (9.0-10.0): {stats['critical']}")
    print(f"Высокие (7.0-8.9): {stats['high']}")
    print(f"Средние (4.0-6.9): {stats['medium']}")
    print(f"Низкие (0.1-3.9): {stats['low']}")
    print(f"Ошибки/не найдено: {stats['errors'] + stats['not_found']}")
    if stats['rate_limited'] > 0:
        print(f"Ошибки лимита запросов (403): {stats['rate_limited']}")
    
    main_logger.info(f"Итоговая статистика: {stats}")
    
    if top_cves:
        print("\nТОП-5 САМЫХ КРИТИЧНЫХ УЯЗВИМОСТЕЙ:")
        for cve_id, score, severity in top_cves:
            print(f"  {cve_id}: {score} ({severity})")
        
        main_logger.info("Топ-5 критичных уязвимостей:")
        for cve_id, score, severity in top_cves:
            main_logger.info(f"  {cve_id}: {score}")
    
    print("\nАнализ завершен!")
    print("Подробный лог сохранен в файл: cve_analyzer.log")
    main_logger.info("Анализ успешно завершен")

if __name__ == "__main__":
    main()
